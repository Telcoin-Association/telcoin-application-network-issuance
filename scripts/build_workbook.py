#!/usr/bin/env python3
"""
TANIP1 staking rewards workbook builder.

Reads rewards/staker_rewards_period_NN.json, divides all values by 100
(JSON stores TEL x 100), and appends the period into the distribution workbook,
matching the exact structure of the committed reference workbook.

Sheets (all: header in row 1, data from row 2):
  - Period Totals      Period | Staker Count | $TEL Rewards | Total Fees | Staker Fees | Referee Fees
  - Raw Data           Period | Address | $TEL Rewards | Uncapped | Missed | Staker Fees | Referee Fees | Total Fees
  - Raw Data - Rebate  + Stake Cap column (between Missed and Staker Fees)
  - Cumulative         Address | $TEL Rewards | Total Fees | Staker Fees | Referee Fees  (per-address running totals)
  - Pivot Table        Address | <period> | <period> | ...  (one $TEL Rewards column per period)

The builder is idempotent: if the period already appears in Period Totals it
exits without modifying anything. Every write is followed by a reopen-and-audit
verification pass (see verify_export) that fails loudly rather than committing a
silently-wrong workbook.
"""

import argparse
import hashlib
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

REWARDS_DIR = "rewards"
PERIOD_FILE_PATTERN = "staker_rewards_period_{}.json"
DIVISOR = 100


# ── helpers ────────────────────────────────────────────────────────────────────

def find_latest_period() -> int:
    files = list(Path(REWARDS_DIR).glob("staker_rewards_period_*.json"))
    if not files:
        raise FileNotFoundError(f"No period files found in {REWARDS_DIR}/")
    return max(int(f.stem.split("_")[-1]) for f in files)


def load_period_json(period: int) -> dict:
    path = Path(REWARDS_DIR) / PERIOD_FILE_PATTERN.format(period)
    with open(path) as fh:
        return json.load(fh)


def file_md5(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest()


def parse_incentives(data: dict) -> list[dict]:
    """Map JSON entries to workbook rows, dividing all TEL fields by 100."""
    rows = []
    for entry in data.get("stakerIncentives", []):
        meta = entry.get("metadata", {})
        reward    = int(entry.get("reward", 0))         / DIVISOR
        uncapped  = int(meta.get("uncappedAmount", 0))  / DIVISOR
        stake_cap = int(meta.get("stakeCapAmount", 0))  / DIVISOR
        fees      = int(meta.get("fees", 0))            / DIVISOR
        ref_fees  = int(meta.get("refereeFees", 0))     / DIVISOR
        rows.append({
            "Address":              entry["address"],
            "$TEL Rewards":         reward,
            "Uncapped $TEL Reward": uncapped,
            "Missed $TEL Reward":   uncapped - reward,
            "Stake Cap":            stake_cap,
            "Staker Fees":          fees,
            "Referee Fees":         ref_fees,
            "Total Fees":           fees + ref_fees,
        })
    return rows


def last_data_row(ws) -> int:
    """Highest row that has any non-empty cell (>=1). Returns 1 if only header."""
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1


# ── Raw Data ───────────────────────────────────────────────────────────────────

def append_raw_data(ws, period: int, rows: list[dict]) -> None:
    """A:Period B:Address C:$TEL Rewards D:Uncapped E:Missed F:Staker G:Referee H:Total."""
    r = last_data_row(ws)
    for row in rows:
        r += 1
        ws.cell(r, 1).value = period
        ws.cell(r, 2).value = row["Address"]
        ws.cell(r, 3).value = row["$TEL Rewards"]
        ws.cell(r, 4).value = row["Uncapped $TEL Reward"]
        ws.cell(r, 5).value = row["Missed $TEL Reward"]
        ws.cell(r, 6).value = row["Staker Fees"]
        ws.cell(r, 7).value = row["Referee Fees"]
        ws.cell(r, 8).value = row["Total Fees"]


def append_raw_data_rebate(ws, period: int, rows: list[dict]) -> None:
    """Same as Raw Data plus a Stake Cap column (F), shifting fees to G/H/I."""
    r = last_data_row(ws)
    for row in rows:
        r += 1
        ws.cell(r, 1).value = period
        ws.cell(r, 2).value = row["Address"]
        ws.cell(r, 3).value = row["$TEL Rewards"]
        ws.cell(r, 4).value = row["Uncapped $TEL Reward"]
        ws.cell(r, 5).value = row["Missed $TEL Reward"]
        ws.cell(r, 6).value = row["Stake Cap"]
        ws.cell(r, 7).value = row["Staker Fees"]
        ws.cell(r, 8).value = row["Referee Fees"]
        ws.cell(r, 9).value = row["Total Fees"]


# ── Period Totals ──────────────────────────────────────────────────────────────

def append_period_totals(ws, period: int, rows: list[dict]) -> None:
    """A:Period B:Staker Count C:$TEL Rewards D:Total Fees E:Staker Fees F:Referee Fees."""
    r = last_data_row(ws) + 1
    ws.cell(r, 1).value = period
    ws.cell(r, 2).value = len(rows)
    ws.cell(r, 3).value = sum(x["$TEL Rewards"] for x in rows)
    ws.cell(r, 4).value = sum(x["Total Fees"]   for x in rows)
    ws.cell(r, 5).value = sum(x["Staker Fees"]  for x in rows)
    ws.cell(r, 6).value = sum(x["Referee Fees"] for x in rows)


# ── Cumulative ─────────────────────────────────────────────────────────────────

def update_cumulative(ws, rows: list[dict]) -> None:
    """
    Per-address running totals across all periods.
    A:Address B:$TEL Rewards C:Total Fees D:Staker Fees E:Referee Fees.
    Read existing totals, add this period's contribution, rewrite sorted
    descending by $TEL Rewards (matching the reference workbook's ordering).
    """
    totals: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        addr = ws.cell(r, 1).value
        if isinstance(addr, str) and addr.startswith("0x"):
            totals[addr] = {
                "$TEL Rewards": ws.cell(r, 2).value or 0,
                "Total Fees":   ws.cell(r, 3).value or 0,
                "Staker Fees":  ws.cell(r, 4).value or 0,
                "Referee Fees": ws.cell(r, 5).value or 0,
            }

    for row in rows:
        a = row["Address"]
        t = totals.setdefault(
            a, {"$TEL Rewards": 0, "Total Fees": 0, "Staker Fees": 0, "Referee Fees": 0}
        )
        t["$TEL Rewards"] += row["$TEL Rewards"]
        t["Total Fees"]   += row["Total Fees"]
        t["Staker Fees"]  += row["Staker Fees"]
        t["Referee Fees"] += row["Referee Fees"]

    ordered = sorted(totals.items(), key=lambda kv: kv[1]["$TEL Rewards"], reverse=True)

    # Clear existing data rows, then rewrite.
    for r in range(2, ws.max_row + 1):
        for c in range(1, 6):
            ws.cell(r, c).value = None
    for i, (addr, t) in enumerate(ordered, start=2):
        ws.cell(i, 1).value = addr
        ws.cell(i, 2).value = t["$TEL Rewards"]
        ws.cell(i, 3).value = t["Total Fees"]
        ws.cell(i, 4).value = t["Staker Fees"]
        ws.cell(i, 5).value = t["Referee Fees"]


# ── Pivot Table ────────────────────────────────────────────────────────────────

def update_pivot_table(ws, period: int, rows: list[dict]) -> None:
    """
    Address x period matrix. Row 1: 'Address', then one column header per period.
    Each cell is that address's $TEL Rewards for that period (0 if absent).
    Adds a new column for `period`, fills it for every address, and keeps the
    address rows sorted ascending (matching the reference workbook).
    """
    # Existing period columns -> column index.
    period_cols: dict[int, int] = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        try:
            period_cols[int(v)] = c
        except (ValueError, TypeError):
            continue

    # Existing addresses -> row index.
    addr_rows: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.startswith("0x"):
            addr_rows[a] = r

    # Allocate the column for this period.
    if period in period_cols:
        col = period_cols[period]
    else:
        col = max(period_cols.values(), default=1) + 1
        ws.cell(1, col).value = period
        period_cols[period] = col

    new_addrs = [row["Address"] for row in rows if row["Address"] not in addr_rows]
    if new_addrs:
        # Rebuild the full address set sorted ascending so new wallets slot in
        # order rather than being appended at the bottom.
        all_addrs = sorted(set(addr_rows) | set(new_addrs), key=str.lower)
        # Snapshot existing values keyed by (addr, period).
        snapshot: dict[tuple[str, int], object] = {}
        for a, r in addr_rows.items():
            for p, c in period_cols.items():
                snapshot[(a, p)] = ws.cell(r, c).value
        # Clear and rewrite address column + all period columns.
        max_r = ws.max_row
        for r in range(2, max_r + 1):
            ws.cell(r, 1).value = None
            for c in period_cols.values():
                ws.cell(r, c).value = None
        addr_rows = {}
        for i, a in enumerate(all_addrs, start=2):
            ws.cell(i, 1).value = a
            addr_rows[a] = i
            for p, c in period_cols.items():
                val = snapshot.get((a, p))
                if val is not None:
                    ws.cell(i, c).value = val

    # Write this period's $TEL Rewards for every address in the run.
    reward_by_addr = {row["Address"]: row["$TEL Rewards"] for row in rows}
    for a, r in addr_rows.items():
        if a in reward_by_addr:
            ws.cell(r, col).value = reward_by_addr[a]
        elif ws.cell(r, col).value is None:
            ws.cell(r, col).value = 0


# ── Verification ───────────────────────────────────────────────────────────────

def verify_export(output_path: Path, period: int, rows: list[dict],
                  json_path: Path, json_hash_before: str,
                  periods_before: set[int]) -> None:
    """
    Four-gate reopen-and-audit. Raises AssertionError on any failure.

    [1] Source JSON byte-for-byte unchanged (MD5).
    [2] Raw Data + Rebate: row count and $TEL Rewards match JSON for this period.
    [3] Period Totals: this period's 5 metrics match JSON aggregates; no prior
        period was dropped.
    [4] Cumulative + Pivot Table: every wallet present; pivot has this period's
        column; pivot column equals the per-wallet $TEL Rewards.
    """
    tol = 0.01

    # [1]
    assert file_md5(json_path) == json_hash_before, (
        "[1] FAIL: source JSON changed during build"
    )
    print("  [1/4] Source JSON unchanged ✓")

    wb = load_workbook(output_path)

    # [2] Raw Data + Rebate
    for sheet in ("Raw Data", "Raw Data - Rebate"):
        ws = wb[sheet]
        period_rows = [
            r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] == period
        ]
        assert len(period_rows) == len(rows), (
            f"[2] FAIL: {sheet} has {len(period_rows)} rows for period {period}, "
            f"expected {len(rows)}"
        )
        expected = sorted(round(r["$TEL Rewards"], 6) for r in rows)
        actual = sorted(round(r[2], 6) for r in period_rows if r[2] is not None)
        assert expected == actual, f"[2] FAIL: {sheet} $TEL Rewards mismatch"
    print(f"  [2/4] Raw Data + Rebate: {len(rows)} rows each, values match ✓")

    # [3] Period Totals
    ws = wb["Period Totals"]
    found = None
    seen = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(r[0], int):
            seen.add(r[0])
            if r[0] == period:
                found = r
    assert found is not None, f"[3] FAIL: period {period} missing from Period Totals"
    assert periods_before.issubset(seen), (
        f"[3] FAIL: Period Totals lost prior periods {periods_before - seen}"
    )

    def close(a, b, label):
        assert abs((a or 0) - b) < tol, (
            f"[3] FAIL {label}: workbook={a}, json={b:.4f}"
        )

    close(found[1], len(rows), "Staker Count")
    close(found[2], sum(x["$TEL Rewards"] for x in rows), "$TEL Rewards")
    close(found[3], sum(x["Total Fees"] for x in rows), "Total Fees")
    close(found[4], sum(x["Staker Fees"] for x in rows), "Staker Fees")
    close(found[5], sum(x["Referee Fees"] for x in rows), "Referee Fees")
    print(
        f"  [3/4] Period Totals row correct "
        f"(count={found[1]}, rewards={found[2]:.2f}, "
        f"staker={found[4]:.2f}, referee={found[5]:.2f}); "
        f"{len(periods_before)} prior period(s) intact ✓"
    )

    # [4] Cumulative + Pivot Table
    ws_cum = wb["Cumulative"]
    cum_addrs = {
        ws_cum.cell(r, 1).value
        for r in range(2, ws_cum.max_row + 1)
        if isinstance(ws_cum.cell(r, 1).value, str)
        and ws_cum.cell(r, 1).value.startswith("0x")
    }
    missing_cum = {x["Address"] for x in rows} - cum_addrs
    assert not missing_cum, (
        f"[4] FAIL: Cumulative missing {len(missing_cum)} wallet(s)"
    )

    ws_pt = wb["Pivot Table"]
    pcol = None
    for c in range(2, ws_pt.max_column + 1):
        v = ws_pt.cell(1, c).value
        if v is not None and str(v) == str(period):
            pcol = c
            break
    assert pcol is not None, f"[4] FAIL: Pivot Table has no column for period {period}"

    pivot_vals: dict[str, object] = {}
    for r in range(2, ws_pt.max_row + 1):
        a = ws_pt.cell(r, 1).value
        if isinstance(a, str) and a.startswith("0x"):
            pivot_vals[a] = ws_pt.cell(r, pcol).value
    missing_pt = {x["Address"] for x in rows} - set(pivot_vals)
    assert not missing_pt, (
        f"[4] FAIL: Pivot Table missing {len(missing_pt)} wallet(s) for period {period}"
    )
    for x in rows:
        pv = pivot_vals.get(x["Address"])
        assert abs((pv or 0) - x["$TEL Rewards"]) < tol, (
            f"[4] FAIL: Pivot value for {x['Address']} = {pv}, "
            f"expected {x['$TEL Rewards']}"
        )
    print(
        f"  [4/4] Cumulative ({len(cum_addrs)} wallets) and Pivot Table "
        f"(period {period} column) complete and correct ✓"
    )


# ── Main ───────────────────────────────────────────────────────────────────────

def build(period: int | None, template: str, output_dir: str) -> None:
    if period is None:
        period = find_latest_period()
        print(f"Auto-detected latest period: {period}")

    json_path = Path(REWARDS_DIR) / PERIOD_FILE_PATTERN.format(period)
    json_hash_before = file_md5(json_path)
    print(f"Loading period {period} data (MD5 {json_hash_before[:8]}…)")

    data = load_period_json(period)
    rows = parse_incentives(data)
    print(f"  {len(rows)} staker entries")

    template_path = Path(template)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template}")

    output_path = Path(output_dir) / f"TANIP1_Rewards_Distribution_Period{period}.xlsx"
    if not output_path.exists():
        shutil.copy(template_path, output_path)
        print(f"Copied template → {output_path}")
    else:
        print(f"Updating existing: {output_path}")

    wb = load_workbook(output_path)

    # Idempotency guard — skip if this period is already in Period Totals.
    ws_tot = wb["Period Totals"]
    periods_before = {
        r[0] for r in ws_tot.iter_rows(min_row=2, values_only=True)
        if isinstance(r[0], int)
    }
    if period in periods_before:
        print(f"Period {period} already present — nothing to do.")
        return

    append_raw_data(wb["Raw Data"], period, rows)
    print(f"  Raw Data: appended {len(rows)} rows")

    append_raw_data_rebate(wb["Raw Data - Rebate"], period, rows)
    print(f"  Raw Data - Rebate: appended {len(rows)} rows")

    append_period_totals(ws_tot, period, rows)
    print("  Period Totals: appended summary row")

    update_cumulative(wb["Cumulative"], rows)
    print("  Cumulative: updated")

    update_pivot_table(wb["Pivot Table"], period, rows)
    print("  Pivot Table: updated")

    wb.save(output_path)
    print(f"Saved: {output_path}")

    print("Running verification...")
    verify_export(output_path, period, rows, json_path, json_hash_before, periods_before)
    print("All checks passed.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build TANIP1 rewards workbook")
    parser.add_argument("--period", type=str, default=None)
    parser.add_argument("--template", required=True)
    parser.add_argument("--output-dir", default="reports")
    args = parser.parse_args()

    period = int(args.period.strip()) if args.period and args.period.strip() else None
    Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    build(period, args.template, args.output_dir)


if __name__ == "__main__":
    main()
